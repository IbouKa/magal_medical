import io
import json
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from sqlalchemy import func
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import (
    User, District, EPS, Affection, Edition,
    FicheJournaliere, LigneConsultation, ServiceDiagnostic, Completude,
    ImportLog, ImportLogLigne
)
from extensions import db

admin_bp = Blueprint('admin', __name__)
PERIODES = ['J-2', 'J-1', 'J', 'J+1', 'J+2', 'J+3']

# Services d'aide au diagnostic — source : feuille "scevice_d'aide" du maquette Excel
# Structure : groupes (LABO, RADIO, ECHO, Scanner, Bloc) avec sous-champs
SERVICES_AIDE = [
    {
        'nom': 'LABO',
        'label': 'Laboratoire',
        'champs': [
            ('labo_malades', 'Nbr de malades'),
            ('labo_examens', "Nbre d'examens"),
            ('labo_poches',  "Poches à sang dist."),
        ]
    },
    {
        'nom': 'RADIO',
        'label': 'Radiologie',
        'champs': [
            ('radio_malades', 'Nbr de malades'),
            ('radio_films',   'Nbre de films'),
            ('radio_examens', "Nbre d'examens"),
        ]
    },
    {
        'nom': 'ECHO',
        'label': 'Echographie',
        'champs': [
            ('echo_malades', 'Nbr de malades'),
            ('echo_films',   'Nbre de films'),
        ]
    },
    {
        'nom': 'SCANNER',
        'label': 'Scanner',
        'champs': [
            ('scanner_malades', 'Nbr de malades'),
            ('scanner_films',   'Nbre de films'),
        ]
    },
    {
        'nom': 'BLOC',
        'label': 'Bloc opératoire',
        'champs': [
            ('bloc', 'Nbr actes'),
        ]
    },
]


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Acces reserve aux administrateurs.", "danger")
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


def get_edition_active():
    return Edition.query.filter_by(active=True).order_by(Edition.annee.desc()).first()


@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    edition = get_edition_active()
    if not edition:
        return render_template('admin/dashboard.html', edition=None, stats={}, periodes=PERIODES)

    total_q = db.session.query(
        func.sum(LigneConsultation.cas_simples),
        func.sum(LigneConsultation.hospitalises),
        func.sum(LigneConsultation.evacues),
        func.sum(LigneConsultation.decedes)
    ).join(FicheJournaliere).filter(FicheJournaliere.edition_id == edition.id).first()

    total_simples = total_q[0] or 0
    total_hospit = total_q[1] or 0
    total_evacues = total_q[2] or 0
    total_decedes = total_q[3] or 0
    total_consultants = total_simples + total_hospit + total_evacues

    nb_eps = EPS.query.filter_by(actif=True).count()
    completude = []
    for p in PERIODES:
        nb = FicheJournaliere.query.filter_by(edition_id=edition.id, periode=p).count()
        completude.append({
            'periode': p,
            'nb_soumis': nb,
            'nb_attendus': nb_eps,
            'pct': round(nb / nb_eps * 100, 1) if nb_eps > 0 else 0
        })

    top_aff = db.session.query(
        Affection.libelle,
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).label('total')
    ).join(LigneConsultation).join(FicheJournaliere).filter(
        FicheJournaliere.edition_id == edition.id
    ).group_by(Affection.libelle).order_by(
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).desc()
    ).limit(10).all()

    fiches_recentes = FicheJournaliere.query.filter_by(
        edition_id=edition.id
    ).order_by(FicheJournaliere.date_saisie.desc()).limit(10).all()

    by_district = db.session.query(
        District.nom,
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).label('total')
    ).join(EPS, District.id == EPS.district_id).join(
        FicheJournaliere, EPS.id == FicheJournaliere.eps_id
    ).join(LigneConsultation).filter(
        FicheJournaliere.edition_id == edition.id
    ).group_by(District.nom).order_by(
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).desc()
    ).all()

    mpe = db.session.query(
        Affection.libelle,
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).label('total')
    ).join(LigneConsultation).join(FicheJournaliere).filter(
        FicheJournaliere.edition_id == edition.id,
        Affection.is_mpe == True
    ).group_by(Affection.libelle).all()

    stats = {
        'total_consultants': total_consultants,
        'total_simples': total_simples,
        'total_hospit': total_hospit,
        'total_evacues': total_evacues,
        'total_decedes': total_decedes,
        'completude': completude,
        'top_affections': top_aff,
        'by_district': by_district,
        'mpe': mpe,
        'fiches_recentes': fiches_recentes,
        'nb_eps': nb_eps,
        'nb_users': User.query.filter_by(actif=True).count(),
    }

    return render_template('admin/dashboard.html', edition=edition, stats=stats, periodes=PERIODES)


@admin_bp.route('/eps')
@login_required
@admin_required
def liste_eps():
    districts = District.query.order_by(District.ordre).all()
    eps_list = EPS.query.join(District).order_by(District.ordre, EPS.ordre).all()
    return render_template('admin/eps.html', eps_list=eps_list, districts=districts)


@admin_bp.route('/eps/nouveau', methods=['GET', 'POST'])
@login_required
@admin_required
def nouveau_eps():
    districts = District.query.order_by(District.ordre).all()
    if request.method == 'POST':
        nom = request.form.get('nom', '').strip().upper()
        type_eps = request.form.get('type_eps', 'poste_sante')
        district_id = request.form.get('district_id')
        responsable = request.form.get('responsable', '').strip()
        telephone = request.form.get('telephone', '').strip()
        code = request.form.get('code', '').strip().upper()
        if not nom or not district_id:
            flash('Nom et district sont obligatoires.', 'danger')
        elif EPS.query.filter_by(nom=nom).first():
            flash('Un EPS avec ce nom existe deja.', 'warning')
        else:
            eps = EPS(nom=nom, type_eps=type_eps, district_id=int(district_id),
                      responsable=responsable, telephone=telephone, code=code or None)
            db.session.add(eps)
            db.session.commit()
            flash('EPS cree avec succes.', 'success')
            return redirect(url_for('admin.liste_eps'))
    return render_template('admin/eps_form.html', eps=None, districts=districts, action='Creer')


@admin_bp.route('/eps/<int:eps_id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_eps(eps_id):
    eps = db.session.get(EPS, eps_id)
    if not eps:
        flash('EPS introuvable.', 'danger')
        return redirect(url_for('admin.liste_eps'))
    districts = District.query.order_by(District.ordre).all()
    if request.method == 'POST':
        eps.nom = request.form.get('nom', '').strip().upper()
        eps.type_eps = request.form.get('type_eps', 'poste_sante')
        eps.district_id = int(request.form.get('district_id'))
        eps.responsable = request.form.get('responsable', '').strip()
        eps.telephone = request.form.get('telephone', '').strip()
        eps.code = request.form.get('code', '').strip().upper() or None
        eps.actif = bool(request.form.get('actif'))
        db.session.commit()
        flash('EPS modifie.', 'success')
        return redirect(url_for('admin.liste_eps'))
    return render_template('admin/eps_form.html', eps=eps, districts=districts, action='Modifier')


@admin_bp.route('/utilisateurs')
@login_required
@admin_required
def liste_utilisateurs():
    users = User.query.order_by(User.role, User.username).all()
    return render_template('admin/utilisateurs.html', users=users)


@admin_bp.route('/utilisateurs/nouveau', methods=['GET', 'POST'])
@login_required
@admin_required
def nouveau_utilisateur():
    eps_list = EPS.query.filter_by(actif=True).order_by(EPS.nom).all()
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        email = request.form.get('email', '').strip() or None
        password = request.form.get('password', '')
        role = request.form.get('role', 'responsable')
        nom_complet = request.form.get('nom_complet', '').strip()
        telephone = request.form.get('telephone', '').strip()
        eps_id = request.form.get('eps_id') or None
        if not username or not password:
            flash('Identifiant et mot de passe obligatoires.', 'danger')
        elif User.query.filter_by(username=username).first():
            flash("Identifiant deja utilise.", 'warning')
        elif len(password) < 6:
            flash('Mot de passe trop court (min 6 caracteres).', 'danger')
        else:
            user = User(username=username, email=email, role=role,
                        nom_complet=nom_complet, telephone=telephone,
                        eps_id=int(eps_id) if eps_id else None)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash('Utilisateur cree.', 'success')
            return redirect(url_for('admin.liste_utilisateurs'))
    return render_template('admin/user_form.html', user=None, eps_list=eps_list, action='Creer')


@admin_bp.route('/utilisateurs/<int:user_id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_utilisateur(user_id):
    user = db.session.get(User, user_id)
    if not user:
        flash('Utilisateur introuvable.', 'danger')
        return redirect(url_for('admin.liste_utilisateurs'))
    eps_list = EPS.query.filter_by(actif=True).order_by(EPS.nom).all()
    if request.method == 'POST':
        user.email = request.form.get('email', '').strip() or None
        user.role = request.form.get('role', 'responsable')
        user.nom_complet = request.form.get('nom_complet', '').strip()
        user.telephone = request.form.get('telephone', '').strip()
        eps_id = request.form.get('eps_id')
        user.eps_id = int(eps_id) if eps_id else None
        user.actif = bool(request.form.get('actif'))
        new_pw = request.form.get('new_password', '').strip()
        if new_pw:
            if len(new_pw) < 6:
                flash('Mot de passe trop court.', 'danger')
                return render_template('admin/user_form.html', user=user, eps_list=eps_list, action='Modifier')
            user.set_password(new_pw)
        db.session.commit()
        flash('Utilisateur modifie.', 'success')
        return redirect(url_for('admin.liste_utilisateurs'))
    return render_template('admin/user_form.html', user=user, eps_list=eps_list, action='Modifier')


@admin_bp.route('/editions')
@login_required
@admin_required
def liste_editions():
    editions = Edition.query.order_by(Edition.annee.desc()).all()
    return render_template('admin/editions.html', editions=editions)


@admin_bp.route('/editions/nouvelle', methods=['GET', 'POST'])
@login_required
@admin_required
def nouvelle_edition():
    if request.method == 'POST':
        annee_str = request.form.get('annee', '')
        libelle = request.form.get('libelle', '').strip()
        active = bool(request.form.get('active'))
        if not annee_str:
            flash("L'annee est obligatoire.", 'danger')
        elif Edition.query.filter_by(annee=int(annee_str)).first():
            flash("Cette edition existe deja.", 'warning')
        else:
            if active:
                Edition.query.update({'active': False})
            ed = Edition(annee=int(annee_str), libelle=libelle, active=active)
            db.session.add(ed)
            db.session.commit()
            flash('Edition creee.', 'success')
            return redirect(url_for('admin.liste_editions'))
    return render_template('admin/edition_form.html', edition=None)


@admin_bp.route('/editions/<int:ed_id>/activer', methods=['POST'])
@login_required
@admin_required
def activer_edition(ed_id):
    Edition.query.update({'active': False})
    ed = db.session.get(Edition, ed_id)
    if ed:
        ed.active = True
        db.session.commit()
        flash(f'Edition {ed.annee} activee.', 'success')
    return redirect(url_for('admin.liste_editions'))


@admin_bp.route('/rapports')
@login_required
@admin_required
def rapports():
    edition = get_edition_active()
    selected_periode = request.args.get('periode', '')
    selected_district = request.args.get('district', '')
    districts = District.query.order_by(District.ordre).all()

    query = db.session.query(
        Affection.numero,
        Affection.libelle,
        Affection.categorie,
        func.sum(LigneConsultation.cas_simples).label('simples'),
        func.sum(LigneConsultation.hospitalises).label('hospit'),
        func.sum(LigneConsultation.evacues).label('evacues'),
        func.sum(LigneConsultation.decedes).label('decedes'),
    ).join(LigneConsultation).join(FicheJournaliere)

    if edition:
        query = query.filter(FicheJournaliere.edition_id == edition.id)
    if selected_periode:
        query = query.filter(FicheJournaliere.periode == selected_periode)
    if selected_district:
        query = query.join(EPS, FicheJournaliere.eps_id == EPS.id).filter(
            EPS.district_id == int(selected_district)
        )

    data = query.group_by(
        Affection.numero, Affection.libelle, Affection.categorie
    ).order_by(Affection.numero).all()

    rapport = []
    for row in data:
        total = (row.simples or 0) + (row.hospit or 0) + (row.evacues or 0)
        rapport.append({
            'numero': row.numero,
            'libelle': row.libelle,
            'categorie': row.categorie,
            'simples': row.simples or 0,
            'hospit': row.hospit or 0,
            'evacues': row.evacues or 0,
            'decedes': row.decedes or 0,
            'total': total
        })

    return render_template('admin/rapports.html',
                           edition=edition,
                           rapport=rapport,
                           periodes=PERIODES,
                           districts=districts,
                           selected_periode=selected_periode,
                           selected_district=selected_district)


@admin_bp.route('/completude')
@login_required
@admin_required
def completude():
    edition = get_edition_active()
    selected_periode = request.args.get('periode', PERIODES[0])
    eps_list = EPS.query.filter_by(actif=True).join(District).order_by(District.ordre, EPS.ordre).all()

    soumis = set()
    if edition:
        fiches = FicheJournaliere.query.filter_by(
            edition_id=edition.id, periode=selected_periode
        ).all()
        soumis = {f.eps_id for f in fiches}

    data = []
    for eps in eps_list:
        data.append({
            'eps': eps,
            'soumis': eps.id in soumis,
        })

    nb_soumis = len(soumis)
    nb_total = len(eps_list)
    pct = round(nb_soumis / nb_total * 100, 1) if nb_total > 0 else 0

    return render_template('admin/completude.html',
                           edition=edition,
                           data=data,
                           periodes=PERIODES,
                           selected_periode=selected_periode,
                           nb_soumis=nb_soumis,
                           nb_total=nb_total,
                           pct=pct)


@admin_bp.route('/saisie-directe/<int:eps_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def saisie_directe(eps_id):
    """L'admin peut saisir/modifier les données de n'importe quel EPS."""
    edition = get_edition_active()
    eps = db.session.get(EPS, eps_id)
    if not eps or not edition:
        flash('EPS ou edition introuvable.', 'danger')
        return redirect(url_for('admin.completude'))

    periode = request.args.get('periode', PERIODES[2])
    affections = Affection.query.filter_by(actif=True).order_by(Affection.numero).all()

    fiche = FicheJournaliere.query.filter_by(
        eps_id=eps_id, edition_id=edition.id, periode=periode
    ).first()

    lignes_map = {}
    if fiche:
        for l in fiche.lignes:
            lignes_map[l.affection_id] = l

    # Charger les données des services d'aide existants
    services_diag_map = {}
    if eps.type_eps == 'hopital':
        sd_existing = ServiceDiagnostic.query.filter_by(
            eps_id=eps_id, edition_id=edition.id, periode=periode
        ).first()
        if sd_existing:
            services_diag_map = {
                'labo_malades':    sd_existing.labo_nb_malades       or 0,
                'labo_examens':    sd_existing.labo_nb_examens       or 0,
                'labo_poches':     sd_existing.labo_poches_sang      or 0,
                'radio_malades':   sd_existing.radio_nb_malades      or 0,
                'radio_films':     sd_existing.radio_nb_films        or 0,
                'radio_examens':   sd_existing.radio_nb_examens      or 0,
                'echo_malades':    sd_existing.echo_nb_malades       or 0,
                'echo_films':      sd_existing.echo_nb_films         or 0,
                'scanner_malades': sd_existing.scanner_nb_malades    or 0,
                'scanner_films':   sd_existing.scanner_nb_films      or 0,
                'bloc':            sd_existing.bloc_nb_interventions or 0,
            }

    if request.method == 'POST':
        periode = request.form.get('periode', periode)
        observations = request.form.get('observations', '')

        fiche = FicheJournaliere.query.filter_by(
            eps_id=eps_id, edition_id=edition.id, periode=periode
        ).first()
        if not fiche:
            fiche = FicheJournaliere(
                eps_id=eps_id, edition_id=edition.id,
                periode=periode, saisi_par=current_user.nom_complet or current_user.username
            )
            db.session.add(fiche)
            db.session.flush()

        fiche.observations = observations
        fiche.date_saisie = datetime.utcnow()

        for aff in affections:
            simples = int(request.form.get(f'simples_{aff.id}', 0) or 0)
            hospit = int(request.form.get(f'hospit_{aff.id}', 0) or 0)
            evacues = int(request.form.get(f'evacues_{aff.id}', 0) or 0)
            decedes = int(request.form.get(f'decedes_{aff.id}', 0) or 0)

            ligne = LigneConsultation.query.filter_by(
                fiche_id=fiche.id, affection_id=aff.id
            ).first()
            if not ligne:
                ligne = LigneConsultation(fiche_id=fiche.id, affection_id=aff.id)
                db.session.add(ligne)
            ligne.cas_simples = simples
            ligne.hospitalises = hospit
            ligne.evacues = evacues
            ligne.decedes = decedes

        # Sauvegarder les services d'aide (hôpitaux uniquement)
        if eps.type_eps == 'hopital':
            sd = ServiceDiagnostic.query.filter_by(
                eps_id=eps_id, edition_id=edition.id, periode=periode
            ).first()
            if not sd:
                sd = ServiceDiagnostic(
                    eps_id=eps_id,
                    edition_id=edition.id,
                    periode=periode
                )
                db.session.add(sd)
            sd.labo_nb_malades       = max(0, int(request.form.get('svc_labo_malades',    0) or 0))
            sd.labo_nb_examens       = max(0, int(request.form.get('svc_labo_examens',    0) or 0))
            sd.labo_poches_sang      = max(0, int(request.form.get('svc_labo_poches',     0) or 0))
            sd.radio_nb_malades      = max(0, int(request.form.get('svc_radio_malades',   0) or 0))
            sd.radio_nb_films        = max(0, int(request.form.get('svc_radio_films',     0) or 0))
            sd.radio_nb_examens      = max(0, int(request.form.get('svc_radio_examens',   0) or 0))
            sd.echo_nb_malades       = max(0, int(request.form.get('svc_echo_malades',    0) or 0))
            sd.echo_nb_films         = max(0, int(request.form.get('svc_echo_films',      0) or 0))
            sd.scanner_nb_malades    = max(0, int(request.form.get('svc_scanner_malades', 0) or 0))
            sd.scanner_nb_films      = max(0, int(request.form.get('svc_scanner_films',   0) or 0))
            sd.bloc_nb_interventions = max(0, int(request.form.get('svc_bloc',            0) or 0))
            sd.date_saisie           = datetime.utcnow()

        db.session.commit()
        flash('Donnees enregistrees.', 'success')
        return redirect(url_for('admin.completude', periode=periode))

    return render_template('admin/saisie_directe.html',
                           eps=eps, edition=edition, fiche=fiche,
                           affections=affections, lignes_map=lignes_map,
                           periodes=PERIODES, selected_periode=periode,
                           services_aide=SERVICES_AIDE,
                           services_diag_map=services_diag_map)


@admin_bp.route('/api/completude-data')
@login_required
@admin_required
def api_completude_data():
    edition = get_edition_active()
    if not edition:
        return jsonify({})
    result = {}
    nb_eps = EPS.query.filter_by(actif=True).count()
    for p in PERIODES:
        nb = FicheJournaliere.query.filter_by(edition_id=edition.id, periode=p).count()
        result[p] = {'nb': nb, 'total': nb_eps,
                     'pct': round(nb / nb_eps * 100, 1) if nb_eps > 0 else 0}
    return jsonify(result)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL — Rapport consolidé (format plat)
# Colonnes : N° | District | Période | Consultants | Cas simples |
#            hospitalisés | Evacuées | décédés | Structures pps/CS | Edition
# Respecte les filtres période et district actifs sur la page Rapports
# ─────────────────────────────────────────────────────────────────────────────
@admin_bp.route('/rapports/supprimer-donnees', methods=['POST'])
@login_required
@admin_required
def supprimer_donnees_rapport():
    """Supprime toutes les fiches journalières (et leurs lignes) correspondant aux filtres actifs."""
    edition = get_edition_active()
    if not edition:
        flash("Aucune edition active.", "warning")
        return redirect(url_for('admin.rapports'))

    selected_periode = request.form.get('periode', '')
    selected_district = request.form.get('district', '')

    # Construire la requête avec les mêmes filtres que la vue rapports
    query = FicheJournaliere.query.filter_by(edition_id=edition.id)

    if selected_periode:
        query = query.filter(FicheJournaliere.periode == selected_periode)
    if selected_district:
        query = query.join(EPS, FicheJournaliere.eps_id == EPS.id).filter(
            EPS.district_id == int(selected_district)
        )

    fiches = query.all()
    nb = len(fiches)

    for fiche in fiches:
        db.session.delete(fiche)

    db.session.commit()

    if nb > 0:
        flash(f"{nb} fiche(s) supprimee(s) avec toutes leurs lignes de consultation.", "success")
    else:
        flash("Aucune fiche trouvee pour ces filtres.", "info")

    return redirect(url_for('admin.rapports',
                            periode=selected_periode,
                            district=selected_district))


@admin_bp.route('/rapports/export-excel')
@login_required
@admin_required
def export_rapport_excel():
    edition = get_edition_active()
    if not edition:
        flash("Aucune edition active.", "warning")
        return redirect(url_for('admin.rapports'))

    selected_periode = request.args.get('periode', '')
    selected_district = request.args.get('district', '')

    # ── Requête : une ligne par (fiche × affection) ───────────
    q = db.session.query(
        District.nom.label('district'),
        FicheJournaliere.periode,
        Affection.libelle,
        LigneConsultation.cas_simples,
        LigneConsultation.hospitalises,
        LigneConsultation.evacues,
        LigneConsultation.decedes,
        EPS.nom.label('eps_nom'),
    ).select_from(LigneConsultation)\
     .join(FicheJournaliere, LigneConsultation.fiche_id == FicheJournaliere.id)\
     .join(Affection, LigneConsultation.affection_id == Affection.id)\
     .join(EPS, FicheJournaliere.eps_id == EPS.id)\
     .join(District, EPS.district_id == District.id)\
     .filter(FicheJournaliere.edition_id == edition.id)

    if selected_periode:
        q = q.filter(FicheJournaliere.periode == selected_periode)
    if selected_district:
        q = q.filter(EPS.district_id == int(selected_district))

    q = q.order_by(District.nom, EPS.nom, FicheJournaliere.periode, Affection.numero)
    rows = q.all()

    if not rows:
        flash("Aucune donnee a exporter avec ces filtres.", "info")
        return redirect(url_for('admin.rapports',
                                periode=selected_periode,
                                district=selected_district))

    # ── Styles ───────────────────────────────────────────────
    hdr_font   = Font(bold=True, size=10, color='1F3D1F')
    hdr_fill   = PatternFill('solid', fgColor='D5E8D4')
    thin       = Side(style='thin', color='AAAAAA')
    brd        = Border(left=thin, right=thin, top=thin, bottom=thin)
    ac         = Alignment(horizontal='center', vertical='center')
    al         = Alignment(horizontal='left',   vertical='center')
    fill_alt   = PatternFill('solid', fgColor='F5F5F5')
    fill_white = PatternFill('solid', fgColor='FFFFFF')

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Rapport'

    # ── En-têtes ─────────────────────────────────────────────
    headers = ['N°', 'District', 'Période', 'Consultants',
               'Cas simples', 'hospitalisés', 'Evacuées', 'décédés',
               'Structures pps/CS', 'Edition']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.border    = brd
        c.alignment = ac
    ws.row_dimensions[1].height = 20

    # ── Données ──────────────────────────────────────────────
    for row_idx, r in enumerate(rows, 2):
        cas_simples  = r.cas_simples  or 0
        hospitalises = r.hospitalises or 0
        evacues      = r.evacues      or 0
        decedes      = r.decedes      or 0

        row_fill = fill_alt if row_idx % 2 == 0 else fill_white
        values   = [row_idx - 1, r.district, r.periode, r.libelle,
                    cas_simples, hospitalises, evacues, decedes,
                    r.eps_nom, edition.annee]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = brd
            c.fill   = row_fill

            if col in (1, 3, 5, 6, 7, 8, 10):
                c.alignment = ac
            else:
                c.alignment = al

            if col == 4:
                c.font = Font(size=10, bold=True)
            elif col == 8 and decedes > 0:
                c.font = Font(size=10, bold=True, color='CC0000')
            else:
                c.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 16

    # ── Largeurs de colonnes ──────────────────────────────────
    col_widths = [6, 22, 10, 44, 13, 15, 13, 12, 26, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A2'

    # ── Génération du fichier ─────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    suffix = ''
    if selected_periode:
        suffix += f'_{selected_periode}'
    if selected_district:
        d = District.query.get(int(selected_district))
        suffix += f'_{d.nom.replace(" ", "_")}' if d else f'_dist{selected_district}'

    filename = f'rapport_magal_{edition.annee}{suffix}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ─────────────────────────────────────────────────────────────────────────────
# TOUTES LES ÉDITIONS — Vue comparée multi-éditions + import Excel
# ─────────────────────────────────────────────────────────────────────────────

# Corrections de noms d'EPS avec accents ou variantes orthographiques
# clé = valeur brute en MAJUSCULES telle qu'elle apparaît dans les exports Excel
# valeur = nom normalisé correspondant à la BDD
EPS_NOM_CORRECTIONS = {
    'PS MBOUSSOBÉ':  'PS MBOUSSOBE',
    'PS HÉLIPORT':   'PS HELIPORT',
    'PS THIAWÈNE':   'PS THIAWENE',
    'PS GUÉDÉ KAW':  'PS GUEDE KAW',
}

# Corrections de libellés d'affections avec variantes orthographiques ou doubles espaces
# clé = libellé brut tel qu'il apparaît dans l'export (sensible à la casse après .strip())
# valeur = libellé normalisé correspondant à la BDD
AFFECTION_CORRECTIONS = {
    'accident circulation Charrette':            'Accidents circulation Charrette',
    'Accident voie Publique':                    'Accidents voie Publique',
    'Cas de Paludisme confirmé':                 'Cas confirmés Paludisme',
    'Affections  Respiratoires':                 'Affections Respiratoires',
    'Affection buco dentaires':                  'Affections buco dentaires',
    'Affections de l\'oeil et annexes':          'Affections de l\'œil et annexes',
    'Affection dermatologiques':                 'Affections dermatologiques',
    'PFA':                                       'Cas suspects PFA',
    'Rougeole':                                  'Cas suspects Rougeole',
    'Méningite':                                 'Cas suspects Méningite',
    'Ictère fébrile':                            'Cas suspects Ictère fébrile',
    'Choléra':                                   'Cas suspects Choléra',
    'Dengue':                                    'Cas suspects Dengue',
    'Fièvres Hémorragiques':                     'Cas suspects Fièvres Hémorragiques',
    'Maladies chroniques (diabète,…)':           'Maladies chroniques (diabète...)',
    
}

def _build_editions_stats():
    """Construit les données statistiques pour toutes les éditions."""
    editions = Edition.query.order_by(Edition.annee).all()

    # ── Résumé par édition ──────────────────────────────────
    stats_by_edition = []
    for ed in editions:
        q = db.session.query(
            func.sum(LigneConsultation.cas_simples),
            func.sum(LigneConsultation.hospitalises),
            func.sum(LigneConsultation.evacues),
            func.sum(LigneConsultation.decedes)
        ).join(FicheJournaliere).filter(FicheJournaliere.edition_id == ed.id).first()
        s = q[0] or 0
        h = q[1] or 0
        e = q[2] or 0
        d = q[3] or 0
        stats_by_edition.append({
            'annee': ed.annee,
            'libelle': ed.libelle or f'Edition {ed.annee}',
            'active': ed.active,
            'consultants': s + h + e,
            'simples': s,
            'hospit': h,
            'evacues': e,
            'decedes': d,
            'nb_fiches': FicheJournaliere.query.filter_by(edition_id=ed.id).count(),
        })

    annees = [s['annee'] for s in stats_by_edition]

    # ── Chart 1 : Consultations totales par édition (barres groupées) ──
    chart_global = {
        'labels': annees,
        'datasets': [
            {'label': 'Cas simples',    'data': [s['simples']     for s in stats_by_edition], 'color': '#198754'},
            {'label': 'Hospitalisés',   'data': [s['hospit']      for s in stats_by_edition], 'color': '#fd7e14'},
            {'label': 'Évacués',        'data': [s['evacues']     for s in stats_by_edition], 'color': '#0dcaf0'},
            {'label': 'Décédés',        'data': [s['decedes']     for s in stats_by_edition], 'color': '#dc3545'},
        ]
    }

    # ── Chart 2 : Évolution par période pour chaque édition (lignes) ──
    palette = ['#198754', '#0d6efd', '#fd7e14', '#dc3545', '#6610f2', '#20c997', '#ffc107', '#0dcaf0']
    period_datasets = []
    for i, ed in enumerate(editions):
        period_data = []
        for p in PERIODES:
            q2 = db.session.query(
                func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues)
            ).join(FicheJournaliere).filter(
                FicheJournaliere.edition_id == ed.id,
                FicheJournaliere.periode == p
            ).scalar() or 0
            period_data.append(q2)
        period_datasets.append({
            'label': str(ed.annee),
            'data': period_data,
            'color': palette[i % len(palette)],
        })
    chart_periode = {
        'labels': PERIODES,
        'datasets': period_datasets,
    }

    # ── Chart 3 : Top 10 affections cumulées toutes éditions ──
    top_aff_rows = db.session.query(
        Affection.libelle,
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).label('total')
    ).join(LigneConsultation).group_by(Affection.libelle).order_by(
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).desc()
    ).limit(10).all()
    chart_top_aff = {
        'labels': [r.libelle for r in top_aff_rows],
        'data':   [r.total   for r in top_aff_rows],
    }

    # ── Chart 4 : Consultations par district cumulées ──
    dist_rows = db.session.query(
        District.nom,
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).label('total')
    ).join(EPS, District.id == EPS.district_id).join(
        FicheJournaliere, EPS.id == FicheJournaliere.eps_id
    ).join(LigneConsultation).group_by(District.nom).order_by(
        func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues).desc()
    ).all()
    chart_district = {
        'labels': [r.nom   for r in dist_rows],
        'data':   [r.total for r in dist_rows],
    }

    # ── Chart 5 : MPE (Maladies à Potentiel Épidémique) par édition ──
    mpe_datasets = []
    mpe_affections = Affection.query.filter_by(is_mpe=True, actif=True).order_by(Affection.numero).all()
    for i, aff in enumerate(mpe_affections):
        mpe_data = []
        for ed in editions:
            val = db.session.query(
                func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues)
            ).join(FicheJournaliere).filter(
                FicheJournaliere.edition_id == ed.id,
                LigneConsultation.affection_id == aff.id
            ).scalar() or 0
            mpe_data.append(val)
        mpe_datasets.append({
            'label': aff.libelle,
            'data':  mpe_data,
            'color': palette[i % len(palette)],
        })
    chart_mpe = {
        'labels': annees,
        'datasets': mpe_datasets,
    }

    return {
        'editions': editions,
        'stats_by_edition': stats_by_edition,
        'chart_global': json.dumps(chart_global),
        'chart_periode': json.dumps(chart_periode),
        'chart_top_aff': json.dumps(chart_top_aff),
        'chart_district': json.dumps(chart_district),
        'chart_mpe': json.dumps(chart_mpe),
    }


@admin_bp.route('/toutes-editions')
@login_required
@admin_required
def toutes_editions():
    data = _build_editions_stats()
    # Import logs (20 derniers, ordre décroissant)
    import_logs = ImportLog.query.order_by(ImportLog.date_import.desc()).limit(20).all()
    # ID du dernier import (pour ouvrir l'accordéon automatiquement après redirect)
    last_import_id = request.args.get('log_id', type=int)
    data['import_logs'] = import_logs
    data['last_import_id'] = last_import_id
    return render_template('admin/toutes_editions.html', **data)


# ─────────────────────────────────────────────────────────────────────────────
# AFFECTIONS — CRUD
# ─────────────────────────────────────────────────────────────────────────────

CATEGORIES_AFFECTION = [
    ('autres',         'Autres'),
    ('paludisme',      'Paludisme'),
    ('mpe',            'MPE — Maladie à Potentiel Épidémique'),
    ('traumatologie',  'Traumatologie'),
    ('dermatologie',   'Dermatologie'),
    ('ophtalmologie',  'Ophtalmologie / ORL'),
    ('gynecologie',    'Gynécologie / Obstétrique'),
    ('chronique',      'Maladies chroniques'),
    ('gastroentero',   'Gastroentérologie'),
    ('respiratoire',   'Affections respiratoires'),
    ('neurologie',     'Neurologie'),
]


@admin_bp.route('/affections')
@login_required
@admin_required
def liste_affections():
    q = request.args.get('q', '').strip()
    cat = request.args.get('categorie', '')
    mpe_filter = request.args.get('mpe', '')

    query = Affection.query
    if q:
        query = query.filter(Affection.libelle.ilike(f'%{q}%'))
    if cat:
        query = query.filter(Affection.categorie == cat)
    if mpe_filter == '1':
        query = query.filter(Affection.is_mpe == True)
    elif mpe_filter == '0':
        query = query.filter(Affection.is_mpe == False)

    affections = query.order_by(Affection.numero).all()
    nb_total = Affection.query.count()
    nb_actifs = Affection.query.filter_by(actif=True).count()
    nb_mpe = Affection.query.filter_by(is_mpe=True).count()

    return render_template(
        'admin/affections.html',
        affections=affections,
        categories=CATEGORIES_AFFECTION,
        q=q,
        selected_categorie=cat,
        selected_mpe=mpe_filter,
        nb_total=nb_total,
        nb_actifs=nb_actifs,
        nb_mpe=nb_mpe,
    )


@admin_bp.route('/affections/nouvelle', methods=['GET', 'POST'])
@login_required
@admin_required
def nouvelle_affection():
    if request.method == 'POST':
        numero_str = request.form.get('numero', '').strip()
        libelle = request.form.get('libelle', '').strip()
        categorie = request.form.get('categorie', 'autres')
        is_mpe = bool(request.form.get('is_mpe'))
        actif = bool(request.form.get('actif'))

        if not numero_str or not libelle:
            flash('Le numéro et le libellé sont obligatoires.', 'danger')
        elif not numero_str.lstrip('-').isdigit():
            flash('Le numéro doit être un entier positif.', 'danger')
        else:
            num = int(numero_str)
            if Affection.query.filter_by(numero=num).first():
                flash(f'Une affection avec le numéro {num} existe déjà.', 'warning')
            elif Affection.query.filter(func.lower(Affection.libelle) == libelle.lower()).first():
                flash('Une affection avec ce libellé existe déjà.', 'warning')
            else:
                aff = Affection(
                    numero=num,
                    libelle=libelle,
                    categorie=categorie,
                    is_mpe=is_mpe,
                    actif=actif,
                )
                db.session.add(aff)
                db.session.commit()
                flash(f'Affection « {libelle} » créée avec succès.', 'success')
                return redirect(url_for('admin.liste_affections'))

    last = Affection.query.order_by(Affection.numero.desc()).first()
    next_num = (last.numero + 1) if last else 1

    return render_template(
        'admin/affection_form.html',
        affection=None,
        categories=CATEGORIES_AFFECTION,
        next_num=next_num,
        action='Créer',
    )


@admin_bp.route('/affections/<int:aff_id>/modifier', methods=['GET', 'POST'])
@login_required
@admin_required
def modifier_affection(aff_id):
    aff = db.session.get(Affection, aff_id)
    if not aff:
        flash('Affection introuvable.', 'danger')
        return redirect(url_for('admin.liste_affections'))

    if request.method == 'POST':
        numero_str = request.form.get('numero', '').strip()
        libelle = request.form.get('libelle', '').strip()
        categorie = request.form.get('categorie', 'autres')
        is_mpe = bool(request.form.get('is_mpe'))
        actif = bool(request.form.get('actif'))

        if not numero_str or not libelle:
            flash('Le numéro et le libellé sont obligatoires.', 'danger')
        elif not numero_str.lstrip('-').isdigit():
            flash('Le numéro doit être un entier positif.', 'danger')
        else:
            num = int(numero_str)
            conflict = Affection.query.filter(
                Affection.numero == num, Affection.id != aff_id
            ).first()
            if conflict:
                flash(f'Le numéro {num} est déjà utilisé par « {conflict.libelle} ».', 'warning')
            else:
                aff.numero = num
                aff.libelle = libelle
                aff.categorie = categorie
                aff.is_mpe = is_mpe
                aff.actif = actif
                db.session.commit()
                flash(f'Affection « {libelle} » modifiée avec succès.', 'success')
                return redirect(url_for('admin.liste_affections'))

    return render_template(
        'admin/affection_form.html',
        affection=aff,
        categories=CATEGORIES_AFFECTION,
        next_num=None,
        action='Modifier',
    )


@admin_bp.route('/affections/<int:aff_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_affection(aff_id):
    aff = db.session.get(Affection, aff_id)
    if aff:
        aff.actif = not aff.actif
        db.session.commit()
        etat = 'activée' if aff.actif else 'désactivée'
        flash(f'Affection « {aff.libelle} » {etat}.', 'success')
    return redirect(url_for('admin.liste_affections'))


@admin_bp.route('/affections/<int:aff_id>/supprimer', methods=['POST'])
@login_required
@admin_required
def supprimer_affection(aff_id):
    aff = db.session.get(Affection, aff_id)
    if not aff:
        flash('Affection introuvable.', 'danger')
        return redirect(url_for('admin.liste_affections'))

    nb_lignes = LigneConsultation.query.filter_by(affection_id=aff_id).count()
    if nb_lignes > 0:
        flash(
            f'Impossible de supprimer « {aff.libelle} » : {nb_lignes} ligne(s) de '
            'consultation liée(s). Désactivez-la plutôt.',
            'warning',
        )
    else:
        libelle = aff.libelle
        db.session.delete(aff)
        db.session.commit()
        flash(f'Affection « {libelle} » supprimée définitivement.', 'success')

    return redirect(url_for('admin.liste_affections'))


@admin_bp.route('/toutes-editions/import-excel', methods=['POST'])
@login_required
@admin_required
def import_excel_editions():
    """
    Importe un fichier Excel au format exporté (10 colonnes) :
    N° | District | Période | Affection | Cas simples | Hospitalisés |
    Evacuées | Décédés | Structures pps/CS | Edition (année)
    Crée les données si elles n'existent pas, les met à jour sinon (upsert).
    Enregistre un log détaillé dans ImportLog / ImportLogLigne.
    """
    file = request.files.get('excel_file')
    if not file or file.filename == '':
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('admin.toutes_editions'))

    # ── Créer le log d'import ──
    import_log = ImportLog(
        filename=file.filename,
        fait_par=current_user.nom_complet or current_user.username,
        statut='ok'
    )
    db.session.add(import_log)
    db.session.flush()   # obtenir import_log.id

    created = 0
    updated = 0
    skipped = 0
    nb_total = 0

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active

        # Cache pour éviter les requêtes répétées
        eps_cache       = {}
        edition_cache   = {}
        affection_cache = {}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Ignorer les lignes vides
            if not row or not any(v is not None for v in row):
                continue

            nb_total += 1

            # ── Extraction des valeurs ──
            periode           = None
            affection_libelle = None
            eps_nom_raw       = None
            edition_annee     = None
            cas_simples = hospitalises = evacues = decedes = 0

            try:
                district_raw      = str(row[1]).strip().upper() if row[1] is not None else None
                periode           = str(row[2]).strip() if row[2] is not None else None
                affection_libelle = str(row[3]).strip() if row[3] is not None else None
                cas_simples       = int(float(row[4])) if row[4] is not None else 0
                hospitalises      = int(float(row[5])) if row[5] is not None else 0
                evacues           = int(float(row[6])) if row[6] is not None else 0
                decedes           = int(float(row[7])) if row[7] is not None else 0
                eps_nom_raw       = str(row[8]).strip().upper() if row[8] is not None else None
                edition_annee     = int(float(row[9])) if row[9] is not None else None
                # Si la colonne Structures pps/CS est vide, utiliser le District comme EPS
                if not eps_nom_raw and district_raw:
                    eps_nom_raw = district_raw
            except (ValueError, TypeError) as ex:
                db.session.add(ImportLogLigne(
                    import_log_id=import_log.id,
                    row_num=row_idx,
                    statut='erreur',
                    eps_nom=eps_nom_raw,
                    periode=periode,
                    affection=affection_libelle,
                    edition_annee=edition_annee,
                    message=f'Valeur invalide : {ex}'
                ))
                skipped += 1
                continue

            # ── Validation champs obligatoires ──
            if not all([periode, affection_libelle, eps_nom_raw, edition_annee]):
                db.session.add(ImportLogLigne(
                    import_log_id=import_log.id,
                    row_num=row_idx,
                    statut='ignore',
                    eps_nom=eps_nom_raw,
                    periode=periode,
                    affection=affection_libelle,
                    edition_annee=edition_annee,
                    message='Champs obligatoires manquants (période, affection, EPS ou édition)'
                ))
                skipped += 1
                continue

            if periode not in PERIODES:
                db.session.add(ImportLogLigne(
                    import_log_id=import_log.id,
                    row_num=row_idx,
                    statut='ignore',
                    eps_nom=eps_nom_raw,
                    periode=periode,
                    affection=affection_libelle,
                    edition_annee=edition_annee,
                    message=f'Période inconnue : « {periode} »'
                ))
                skipped += 1
                continue

            # ── Edition (find or create) ──
            if edition_annee not in edition_cache:
                ed = Edition.query.filter_by(annee=edition_annee).first()
                if not ed:
                    ed = Edition(
                        annee=edition_annee,
                        libelle=f'Grand Magal de Touba Edition {edition_annee}',
                        active=False
                    )
                    db.session.add(ed)
                    db.session.flush()
                edition_cache[edition_annee] = ed
            edition = edition_cache[edition_annee]

            # ── Pré-traitement : correction des noms d'EPS accentués ──
            eps_nom_raw = EPS_NOM_CORRECTIONS.get(eps_nom_raw, eps_nom_raw)

            # ── EPS (find, case-insensitive) ──
            if eps_nom_raw not in eps_cache:
                eps = EPS.query.filter(func.upper(EPS.nom) == eps_nom_raw).first()
                if not eps:
                    eps = EPS.query.filter(EPS.nom.ilike(f'{eps_nom_raw[:15]}%')).first()
                eps_cache[eps_nom_raw] = eps
            eps = eps_cache[eps_nom_raw]

            if not eps:
                db.session.add(ImportLogLigne(
                    import_log_id=import_log.id,
                    row_num=row_idx,
                    statut='ignore',
                    eps_nom=eps_nom_raw,
                    periode=periode,
                    affection=affection_libelle,
                    edition_annee=edition_annee,
                    message=f'EPS introuvable : « {eps_nom_raw} »'
                ))
                skipped += 1
                continue

            # ── Pré-traitement : correction des libellés d'affections ──
            affection_libelle = AFFECTION_CORRECTIONS.get(affection_libelle, affection_libelle)

            # ── Affection ──
            aff_key = affection_libelle.lower()
            if aff_key not in affection_cache:
                aff = Affection.query.filter(func.lower(Affection.libelle) == aff_key).first()
                if not aff:
                    aff = Affection.query.filter(
                        Affection.libelle.ilike(f'{affection_libelle[:20]}%')
                    ).first()
                affection_cache[aff_key] = aff
            affection = affection_cache[aff_key]

            if not affection:
                db.session.add(ImportLogLigne(
                    import_log_id=import_log.id,
                    row_num=row_idx,
                    statut='ignore',
                    eps_nom=eps_nom_raw,
                    periode=periode,
                    affection=affection_libelle[:200],
                    edition_annee=edition_annee,
                    message=f'Affection introuvable : « {affection_libelle[:60]} »'
                ))
                skipped += 1
                continue

            # ── FicheJournaliere (find or create) ──
            fiche = FicheJournaliere.query.filter_by(
                eps_id=eps.id, edition_id=edition.id, periode=periode
            ).first()
            if not fiche:
                fiche = FicheJournaliere(
                    eps_id=eps.id, edition_id=edition.id,
                    periode=periode, saisi_par='Import Excel', statut='soumis'
                )
                db.session.add(fiche)
                db.session.flush()

            # ── LigneConsultation (upsert) ──
            ligne = LigneConsultation.query.filter_by(
                fiche_id=fiche.id, affection_id=affection.id
            ).first()
            if ligne:
                ligne.cas_simples  = max(0, cas_simples)
                ligne.hospitalises = max(0, hospitalises)
                ligne.evacues      = max(0, evacues)
                ligne.decedes      = max(0, decedes)
                statut_ligne = 'mis_a_jour'
                updated += 1
            else:
                ligne = LigneConsultation(
                    fiche_id=fiche.id, affection_id=affection.id,
                    cas_simples=max(0, cas_simples),
                    hospitalises=max(0, hospitalises),
                    evacues=max(0, evacues),
                    decedes=max(0, decedes)
                )
                db.session.add(ligne)
                statut_ligne = 'cree'
                created += 1

            db.session.add(ImportLogLigne(
                import_log_id=import_log.id,
                row_num=row_idx,
                statut=statut_ligne,
                eps_nom=eps.nom,
                periode=periode,
                affection=affection.libelle[:200],
                edition_annee=edition_annee,
                message=None
            ))

        # Mettre à jour les compteurs du log
        import_log.nb_created = created
        import_log.nb_updated = updated
        import_log.nb_skipped = skipped
        import_log.nb_total   = nb_total
        import_log.statut     = 'ok'

        db.session.commit()

        msg = f'Import terminé : {created} ligne(s) créée(s), {updated} mise(s) à jour'
        if skipped:
            msg += f', {skipped} ignorée(s)'
        msg += '.'
        flash(msg, 'success')

    except Exception as ex:
        import_log.statut = 'erreur'
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        flash(f"Erreur lors de l'import : {ex}", 'danger')
        return redirect(url_for('admin.toutes_editions'))

    return redirect(url_for('admin.toutes_editions', log_id=import_log.id) + '#tab-logs')
