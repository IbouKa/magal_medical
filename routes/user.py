import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, abort, send_file
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import (
    EPS, Affection, Edition,
    FicheJournaliere, LigneConsultation, ServiceDiagnostic
)
from extensions import db

user_bp = Blueprint('user', __name__)
PERIODES = ['J-2', 'J-1', 'J', 'J+1', 'J+2', 'J+3']


def get_edition_active():
    return Edition.query.filter_by(active=True).order_by(Edition.annee.desc()).first()


def user_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if not current_user.eps_id and not current_user.is_admin:
            flash("Votre compte n'est pas associe a un EPS.", "warning")
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


@user_bp.route('/dashboard')
@login_required
@user_required
def dashboard():
    edition = get_edition_active()
    eps = current_user.eps
    if not eps:
        flash("Aucun EPS associe a votre compte.", "warning")
        return redirect(url_for('auth.login'))

    fiches = []
    if edition:
        fiches = FicheJournaliere.query.filter_by(
            eps_id=eps.id, edition_id=edition.id
        ).order_by(FicheJournaliere.periode).all()

    periodes_soumises = {f.periode for f in fiches}

    summary = []
    for p in PERIODES:
        fiche = next((f for f in fiches if f.periode == p), None)
        total = fiche.get_total_consultants() if fiche else 0
        decedes = fiche.get_total_decedes() if fiche else 0
        summary.append({
            'periode': p,
            'soumis': p in periodes_soumises,
            'total': total,
            'decedes': decedes,
            'fiche_id': fiche.id if fiche else None
        })

    return render_template('user/dashboard.html',
                           eps=eps, edition=edition,
                           summary=summary,
                           periodes=PERIODES)


@user_bp.route('/saisie', methods=['GET', 'POST'])
@login_required
@user_required
def saisie():
    edition = get_edition_active()
    eps = current_user.eps
    if not eps or not edition:
        flash("Edition inactive ou EPS non configure.", "warning")
        return redirect(url_for('user.dashboard'))

    periode = request.args.get('periode', PERIODES[2])
    if periode not in PERIODES:
        periode = PERIODES[2]

    affections = Affection.query.filter_by(actif=True).order_by(Affection.numero).all()

    fiche = FicheJournaliere.query.filter_by(
        eps_id=eps.id, edition_id=edition.id, periode=periode
    ).first()

    lignes_map = {}
    if fiche:
        for ligne in fiche.lignes:
            lignes_map[ligne.affection_id] = ligne

    if request.method == 'POST':
        periode_post = request.form.get('periode', periode)
        if periode_post not in PERIODES:
            periode_post = PERIODES[2]
        observations = request.form.get('observations', '')

        fiche = FicheJournaliere.query.filter_by(
            eps_id=eps.id, edition_id=edition.id, periode=periode_post
        ).first()

        if not fiche:
            fiche = FicheJournaliere(
                eps_id=eps.id,
                edition_id=edition.id,
                periode=periode_post,
                saisi_par=current_user.nom_complet or current_user.username,
                statut='soumis'
            )
            db.session.add(fiche)
            db.session.flush()
        else:
            fiche.date_saisie = datetime.utcnow()

        fiche.observations = observations

        for aff in affections:
            simples = int(request.form.get(f'simples_{aff.id}', 0) or 0)
            hospit = int(request.form.get(f'hospit_{aff.id}', 0) or 0)
            evacues = int(request.form.get(f'evacues_{aff.id}', 0) or 0)
            decedes = int(request.form.get(f'decedes_{aff.id}', 0) or 0)

            ligne = LigneConsultation.query.filter_by(
                fiche_id=fiche.id, affection_id=aff.id
            ).first()
            if not ligne:
                ligne = LigneConsultation(fiche_id=fiche.id, affection_id=aff.id)
                db.session.add(ligne)

            ligne.cas_simples = max(0, simples)
            ligne.hospitalises = max(0, hospit)
            ligne.evacues = max(0, evacues)
            ligne.decedes = max(0, decedes)

        db.session.commit()
        flash(f'Fiche {periode_post} enregistree avec succes.', 'success')
        return redirect(url_for('user.dashboard'))

    return render_template('user/saisie.html',
                           eps=eps, edition=edition,
                           fiche=fiche, affections=affections,
                           lignes_map=lignes_map,
                           periodes=PERIODES,
                           selected_periode=periode)


@user_bp.route('/mes-fiches')
@login_required
@user_required
def mes_fiches():
    edition = get_edition_active()
    eps = current_user.eps
    fiches = []
    if eps and edition:
        fiches = FicheJournaliere.query.filter_by(
            eps_id=eps.id, edition_id=edition.id
        ).order_by(FicheJournaliere.periode).all()
    return render_template('user/mes_fiches.html', eps=eps, edition=edition, fiches=fiches)


@user_bp.route('/fiche/<int:fiche_id>')
@login_required
@user_required
def voir_fiche(fiche_id):
    fiche = db.session.get(FicheJournaliere, fiche_id)
    if not fiche:
        abort(404)

    # Strict ownership check — un non-admin ne peut voir que les fiches de son EPS
    if not current_user.is_admin and fiche.eps_id != current_user.eps_id:
        abort(403)

    affections = Affection.query.filter_by(actif=True).order_by(Affection.numero).all()
    lignes_map = {l.affection_id: l for l in fiche.lignes}

    return render_template('user/voir_fiche.html',
                           fiche=fiche,
                           affections=affections,
                           lignes_map=lignes_map)


# ─────────────────────────────────────────────────────────────────────────────
# STATISTIQUES — filtrées par EPS (structure de l'utilisateur)
# ─────────────────────────────────────────────────────────────────────────────
@user_bp.route('/mes-stats')
@login_required
@user_required
def mes_stats():
    edition = get_edition_active()
    eps = current_user.eps
    if not eps:
        flash("Aucun EPS associe a votre compte.", "warning")
        return redirect(url_for('user.dashboard'))

    total_simples = total_hospit = total_evacues = total_decedes = 0
    total_consultants = 0
    periode_data = []
    top_affections = []
    mpe_data = []

    if edition:
        # Totaux globaux pour cet EPS
        result = db.session.query(
            func.sum(LigneConsultation.cas_simples),
            func.sum(LigneConsultation.hospitalises),
            func.sum(LigneConsultation.evacues),
            func.sum(LigneConsultation.decedes)
        ).join(FicheJournaliere).filter(
            FicheJournaliere.eps_id == eps.id,
            FicheJournaliere.edition_id == edition.id
        ).first()

        total_simples  = result[0] or 0
        total_hospit   = result[1] or 0
        total_evacues  = result[2] or 0
        total_decedes  = result[3] or 0
        total_consultants = total_simples + total_hospit + total_evacues

        # Données par période pour cet EPS
        for p in PERIODES:
            q = db.session.query(
                func.sum(LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues),
                func.sum(LigneConsultation.decedes)
            ).join(FicheJournaliere).filter(
                FicheJournaliere.eps_id == eps.id,
                FicheJournaliere.edition_id == edition.id,
                FicheJournaliere.periode == p
            ).first()
            total_p   = q[0] or 0
            decedes_p = q[1] or 0
            fiche_p = FicheJournaliere.query.filter_by(
                eps_id=eps.id, edition_id=edition.id, periode=p
            ).first()
            periode_data.append({
                'periode': p,
                'total': total_p,
                'decedes': decedes_p,
                'soumis': fiche_p is not None,
                'fiche_id': fiche_p.id if fiche_p else None
            })

        # Top 10 affections pour cet EPS
        top_affections = db.session.query(
            Affection.libelle,
            func.sum(
                LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues
            ).label('total')
        ).join(LigneConsultation).join(FicheJournaliere).filter(
            FicheJournaliere.eps_id == eps.id,
            FicheJournaliere.edition_id == edition.id
        ).group_by(Affection.libelle).order_by(
            func.sum(
                LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues
            ).desc()
        ).limit(10).all()

        # MPE pour cet EPS
        mpe_data = db.session.query(
            Affection.libelle,
            func.sum(
                LigneConsultation.cas_simples + LigneConsultation.hospitalises + LigneConsultation.evacues
            ).label('total'),
            func.sum(LigneConsultation.decedes).label('decedes')
        ).join(LigneConsultation).join(FicheJournaliere).filter(
            FicheJournaliere.eps_id == eps.id,
            FicheJournaliere.edition_id == edition.id,
            Affection.is_mpe == True
        ).group_by(Affection.libelle).all()

    stats_data = {
        'total_consultants': total_consultants,
        'total_simples': total_simples,
        'total_hospit': total_hospit,
        'total_evacues': total_evacues,
        'total_decedes': total_decedes,
        'top_affections': top_affections,
        'periode_data': periode_data,
        'mpe_data': mpe_data,
    }

    return render_template('user/mes_stats.html',
                           eps=eps,
                           edition=edition,
                           stats=stats_data,
                           periodes=PERIODES)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL — Mes Fiches  (format plat : une ligne par affection × période)
# Colonnes : N° | District | Période | Consultants | Cas simples |
#            hospitalisés | Evacuées | décédés | Structures pps/CS | Edition
# ─────────────────────────────────────────────────────────────────────────────
@user_bp.route('/mes-fiches/export-excel')
@login_required
@user_required
def export_fiches_excel():
    edition = get_edition_active()
    eps = current_user.eps
    if not eps or not edition:
        flash("Aucune edition active ou EPS non configure.", "warning")
        return redirect(url_for('user.mes_fiches'))

    fiches = FicheJournaliere.query.filter_by(
        eps_id=eps.id, edition_id=edition.id
    ).order_by(FicheJournaliere.periode).all()

    if not fiches:
        flash("Aucune fiche a exporter.", "info")
        return redirect(url_for('user.mes_fiches'))

    affections = Affection.query.filter_by(actif=True).order_by(Affection.numero).all()
    district_nom = eps.district.nom if eps.district else ''

    # ── Construction des lignes ───────────────────────────────
    rows = []
    for fiche in fiches:
        lignes_map = {l.affection_id: l for l in fiche.lignes}
        for aff in affections:
            ligne = lignes_map.get(aff.id)
            rows.append((
                district_nom,
                fiche.periode,
                aff.libelle,
                ligne.cas_simples if ligne else 0,
                ligne.hospitalises if ligne else 0,
                ligne.evacues if ligne else 0,
                ligne.decedes if ligne else 0,
                eps.nom,
                edition.annee,
            ))

    # ── Styles ───────────────────────────────────────────────
    hdr_font   = Font(bold=True, size=10, color='1F3D1F')
    hdr_fill   = PatternFill('solid', fgColor='D5E8D4')
    thin       = Side(style='thin', color='AAAAAA')
    brd        = Border(left=thin, right=thin, top=thin, bottom=thin)
    ac         = Alignment(horizontal='center', vertical='center')
    al         = Alignment(horizontal='left',   vertical='center')
    fill_alt   = PatternFill('solid', fgColor='F5F5F5')
    fill_white = PatternFill('solid', fgColor='FFFFFF')

    wb  = Workbook()
    ws  = wb.active
    ws.title = 'Fiches'

    # ── En-têtes ─────────────────────────────────────────────
    headers = ['N°', 'District', 'Période', 'Consultants',
               'Cas simples', 'hospitalisés', 'Evacuées', 'décédés',
               'Structures pps/CS', 'Edition']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font    = hdr_font
        c.fill    = hdr_fill
        c.border  = brd
        c.alignment = ac
    ws.row_dimensions[1].height = 20

    # ── Données ──────────────────────────────────────────────
    for row_idx, (district, periode, affection,
                  cas_simples, hospitalises, evacues, decedes,
                  eps_nom, annee) in enumerate(rows, 2):

        row_fill = fill_alt if row_idx % 2 == 0 else fill_white
        values   = [row_idx - 1, district, periode, affection,
                    cas_simples, hospitalises, evacues, decedes,
                    eps_nom, annee]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = brd
            c.fill   = row_fill

            # Alignement
            if col in (1, 3, 5, 6, 7, 8, 10):
                c.alignment = ac
            else:
                c.alignment = al

            # Police
            if col == 4:                              # affection — gras
                c.font = Font(size=10, bold=True)
            elif col == 8 and decedes and decedes > 0: # décédés — rouge
                c.font = Font(size=10, bold=True, color='CC0000')
            else:
                c.font = Font(size=10)

        ws.row_dimensions[row_idx].height = 16

    # ── Largeurs de colonnes ──────────────────────────────────
    col_widths = [6, 22, 10, 44, 13, 15, 13, 12, 26, 10]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Figer la ligne d'en-tête
    ws.freeze_panes = 'A2'

    # ── Génération du fichier ─────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_nom = eps.nom.replace(' ', '_').replace('/', '-')
    filename = f'fiches_{safe_nom}_{edition.annee}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
